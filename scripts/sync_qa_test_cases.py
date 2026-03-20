#!/usr/bin/env python3
"""Sync QA test cases from the markdown guide into the Excel tracking spreadsheet.

Usage:
    # Check for drift (exits non-zero if out of sync)
    python scripts/sync_qa_test_cases.py --check

    # Update the Excel to match the markdown (preserves Status/Tester/Date/Notes)
    python scripts/sync_qa_test_cases.py --sync

Source of truth: docs/testing/01-qa-testing-guide-baremetal.md
Target:          docs/testing/ecube-qa-test-cases.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import NamedTuple

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    sys.exit(
        "openpyxl is required.  Install it with:\n"
        "  pip install openpyxl\n"
        "Or install the project dev extras:\n"
        "  pip install -e '.[dev]'"
    )

# ---------------------------------------------------------------------------
# Paths (relative to repo root)
# ---------------------------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parent.parent
MD_PATH = REPO_ROOT / "docs" / "testing" / "01-qa-testing-guide-baremetal.md"
XLSX_PATH = REPO_ROOT / "docs" / "testing" / "ecube-qa-test-cases.xlsx"

# ---------------------------------------------------------------------------
# Section-name mapping  (markdown header text → short Excel label)
# ---------------------------------------------------------------------------
SECTION_SHORT_NAMES: dict[str, str] = {
    "Login Endpoint (`POST /auth/token`)": "Login Endpoint",
    "Authorization": "Authorization",
    "Project Isolation": "Project Isolation",
    "Drive State Machine": "Drive State Machine",
    "Filesystem Detection": "Filesystem Detection",
    "Drive Formatting": "Drive Formatting",
    "Port Enablement": "Port Enablement",
    "Hub & Port Identification Enrichment": "Hub/Port Enrichment",
    "USB Hardware (Bare-Metal Specific)": "USB Hardware",
    "End-to-End Copy Workflow": "End-to-End Copy",
    "Error Handling": "Error Handling",
    "User Role Management": "User Role Mgmt",
    "OS User & Group Management": "OS User & Group Mgmt",
    "First-Run Setup": "First-Run Setup",
    "Database Provisioning API": "Database Provisioning",
}

# Excel styling constants
HEADER_FONT = Font(bold=True, size=11)
SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------
class TestCase(NamedTuple):
    """A single test case parsed from the markdown."""
    section_num: str       # e.g. "12.3"
    section_label: str     # Short label for Excel, e.g. "12.3 Project Isolation"
    row_num: int           # 1-based row within section
    tc_id: str             # e.g. "TC-301"
    test_name: str
    steps: str             # "How" or "Steps" column; empty for 3-column tables
    expected: str


class Section(NamedTuple):
    section_num: str
    section_label: str
    header_label: str      # "§12.3 Project Isolation" for the section header row
    test_cases: list[TestCase]


# ---------------------------------------------------------------------------
# Markdown parser
# ---------------------------------------------------------------------------
_SECTION_RE = re.compile(r"^###\s+(12\.(\d+(?:\.\d+)?))\s+(.+)$")
_TABLE_ROW_RE = re.compile(r"^\|\s*(\d+)\s*\|(.+)\|$")
_SEPARATOR_RE = re.compile(r"^\|\s*-+")


def _section_key(section_num: str) -> int:
    """Convert a section number like '12.4.1' to a sortable TC-ID prefix.

    - ``12.4``  → 400
    - ``12.4.1`` → 410
    - ``12.4.2`` → 420
    - ``12.10`` → 1000
    """
    parts = section_num.split(".")
    minor = int(parts[1])          # e.g. 4
    sub = int(parts[2]) if len(parts) > 2 else 0  # e.g. 1
    return minor * 100 + sub * 10


def _short_label(section_num: str, raw_name: str) -> str:
    """Convert a raw markdown section name to the short Excel label."""
    short = SECTION_SHORT_NAMES.get(raw_name, raw_name)
    return f"{section_num} {short}"


def _tc_id(key: int, row_num: int) -> str:
    return f"TC-{key + row_num}"


def _strip_md(text: str) -> str:
    """Strip leading/trailing whitespace and inline backtick formatting."""
    text = text.strip()
    # Remove all backtick characters (inline code formatting)
    text = text.replace("`", "")
    return text


def parse_markdown(path: Path) -> list[Section]:
    """Parse all ### 12.x sections and return structured test cases."""
    lines = path.read_text(encoding="utf-8").splitlines()
    sections: list[Section] = []
    i = 0
    while i < len(lines):
        m = _SECTION_RE.match(lines[i])
        if not m:
            i += 1
            continue

        section_num = m.group(1)         # "12.3" or "12.4.1"
        raw_name = m.group(3).strip()    # "Project Isolation"
        key = _section_key(section_num)
        label = _short_label(section_num, raw_name)
        header_label = f"§{label}"

        # Advance past the section header to find the table
        i += 1
        test_cases: list[TestCase] = []

        # Locate table header row: | # | ... |
        while i < len(lines) and not _SECTION_RE.match(lines[i]):
            if lines[i].strip().startswith("| #"):
                # Determine column count from header
                cols = [c.strip() for c in lines[i].strip().strip("|").split("|")]
                num_cols = len(cols)
                # Skip separator line
                i += 1
                if i < len(lines) and _SEPARATOR_RE.match(lines[i]):
                    i += 1

                # Parse data rows
                while i < len(lines):
                    row_match = _TABLE_ROW_RE.match(lines[i])
                    if not row_match:
                        break
                    row_num = int(row_match.group(1))
                    # Split the rest of the row by |
                    rest = row_match.group(2)
                    parts = rest.split("|")

                    if num_cols == 4:
                        # | # | Test | How/Steps | Expected |
                        test_name = _strip_md(parts[0]) if len(parts) > 0 else ""
                        steps = _strip_md(parts[1]) if len(parts) > 1 else ""
                        expected = _strip_md(parts[2]) if len(parts) > 2 else ""
                    elif num_cols == 3:
                        # | # | Test | Expected |
                        test_name = _strip_md(parts[0]) if len(parts) > 0 else ""
                        steps = ""
                        expected = _strip_md(parts[1]) if len(parts) > 1 else ""
                    else:
                        test_name = _strip_md(parts[0]) if len(parts) > 0 else ""
                        steps = _strip_md("|".join(parts[1:-1])) if len(parts) > 2 else ""
                        expected = _strip_md(parts[-1]) if len(parts) > 0 else ""

                    test_cases.append(TestCase(
                        section_num=section_num,
                        section_label=label,
                        row_num=row_num,
                        tc_id=_tc_id(key, row_num),
                        test_name=test_name,
                        steps=steps,
                        expected=expected,
                    ))
                    i += 1
                break  # done with this section's table
            i += 1

        # Handle sections without tables (e.g. 12.6 End-to-End Copy)
        if not test_cases:
            # Check if the section has narrative content — create a single summary TC
            if "end-to-end" in raw_name.lower() or "e2e" in raw_name.lower():
                test_cases.append(TestCase(
                    section_num=section_num,
                    section_label=label,
                    row_num=1,
                    tc_id=_tc_id(key, 1),
                    test_name="Full E2E workflow",
                    steps=(
                        "Setup test files → add mount → discover drive → "
                        "initialize → create job → start → poll → verify → "
                        "manifest → eject → verify on another machine"
                    ),
                    expected=(
                        "All steps succeed; checksums match; "
                        "audit trail shows complete chain"
                    ),
                ))

        sections.append(Section(
            section_num=section_num,
            section_label=label,
            header_label=header_label,
            test_cases=test_cases,
        ))

    return sections


# ---------------------------------------------------------------------------
# Excel reader — extract existing tracking data
# ---------------------------------------------------------------------------
def _load_existing_tracking(path: Path) -> dict[str, dict[str, str | None]]:
    """Load Status/Tester/Date/Notes by TC ID from the existing spreadsheet."""
    tracking: dict[str, dict[str, str | None]] = {}
    if not path.exists():
        return tracking
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for row in range(2, (ws.max_row or 1) + 1):
        tc_id = ws.cell(row=row, column=1).value
        if tc_id and str(tc_id).startswith("TC-"):
            tracking[str(tc_id)] = {
                "status": ws.cell(row=row, column=6).value,
                "tester": ws.cell(row=row, column=7).value,
                "date": ws.cell(row=row, column=8).value,
                "notes": ws.cell(row=row, column=9).value,
            }
    wb.close()
    return tracking


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------
HEADERS = ["TC ID", "Section", "Test Name", "Steps / How", "Expected Result",
           "Status", "Tester", "Date", "Notes"]
COL_WIDTHS = [12, 28, 30, 55, 45, 10, 12, 12, 25]


def _write_xlsx(sections: list[Section], tracking: dict, path: Path) -> int:
    """Generate the Excel spreadsheet.  Returns total test case count."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QA Test Cases"

    # Write headers
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGNMENT

    # Set column widths
    for col, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    row = 2
    total = 0

    for section in sections:
        # Section header row
        cell = ws.cell(row=row, column=1, value=section.header_label)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        ws.cell(row=row, column=2, value=section.section_label).font = SECTION_FONT
        ws.cell(row=row, column=2).fill = SECTION_FILL
        for col in range(3, len(HEADERS) + 1):
            ws.cell(row=row, column=col).fill = SECTION_FILL
        row += 1

        # Test case rows
        for tc in section.test_cases:
            ws.cell(row=row, column=1, value=tc.tc_id)
            ws.cell(row=row, column=2, value=tc.section_label)
            ws.cell(row=row, column=3, value=tc.test_name)
            ws.cell(row=row, column=4, value=tc.steps)
            ws.cell(row=row, column=5, value=tc.expected)

            # Restore tracking data if it existed
            prev = tracking.get(tc.tc_id, {})
            ws.cell(row=row, column=6, value=prev.get("status"))
            ws.cell(row=row, column=7, value=prev.get("tester"))
            ws.cell(row=row, column=8, value=prev.get("date"))
            ws.cell(row=row, column=9, value=prev.get("notes"))

            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row, column=col).alignment = WRAP_ALIGNMENT

            row += 1
            total += 1

    # Blank row + total
    row += 1
    ws.cell(row=row, column=1, value="Total Test Cases:")
    cell = ws.cell(row=row, column=2, value=total)
    cell.font = HEADER_FONT
    ws.cell(row=row, column=1).font = HEADER_FONT

    wb.save(path)
    return total


# ---------------------------------------------------------------------------
# Diff / check logic
# ---------------------------------------------------------------------------
def _load_xlsx_cases(path: Path) -> dict[str, tuple[str, str, str, str]]:
    """Return {tc_id: (section, test_name, steps, expected)} from Excel."""
    cases: dict[str, tuple[str, str, str, str]] = {}
    if not path.exists():
        return cases
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for row in range(2, (ws.max_row or 1) + 1):
        tc_id = ws.cell(row=row, column=1).value
        if tc_id and str(tc_id).startswith("TC-"):
            cases[str(tc_id)] = (
                str(ws.cell(row=row, column=2).value or ""),
                str(ws.cell(row=row, column=3).value or ""),
                str(ws.cell(row=row, column=4).value or ""),
                str(ws.cell(row=row, column=5).value or ""),
            )
    wb.close()
    return cases


def check_sync(sections: list[Section], xlsx_path: Path) -> list[str]:
    """Compare markdown test cases against the Excel.  Return list of diffs."""
    xlsx_cases = _load_xlsx_cases(xlsx_path)
    md_cases: dict[str, tuple[str, str, str, str]] = {}
    for section in sections:
        for tc in section.test_cases:
            md_cases[tc.tc_id] = (tc.section_label, tc.test_name,
                                  tc.steps, tc.expected)

    diffs: list[str] = []

    # Cases in markdown but not in Excel
    for tc_id in sorted(md_cases.keys() - xlsx_cases.keys()):
        diffs.append(f"  ADD  {tc_id}: {md_cases[tc_id][1]}")

    # Cases in Excel but not in markdown
    for tc_id in sorted(xlsx_cases.keys() - md_cases.keys()):
        diffs.append(f"  DEL  {tc_id}: {xlsx_cases[tc_id][1]}")

    # Cases with different content
    for tc_id in sorted(md_cases.keys() & xlsx_cases.keys()):
        md_vals = md_cases[tc_id]
        xl_vals = xlsx_cases[tc_id]
        if md_vals != xl_vals:
            diffs.append(f"  MOD  {tc_id}: {md_vals[1]}")

    return diffs


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Sync QA test cases: markdown → Excel"
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "--check", action="store_true",
        help="Report drift and exit non-zero if out of sync",
    )
    group.add_argument(
        "--sync", action="store_true",
        help="Regenerate the Excel from the markdown (preserves tracking columns)",
    )
    args = parser.parse_args()

    if not MD_PATH.exists():
        sys.exit(f"Markdown file not found: {MD_PATH}")

    sections = parse_markdown(MD_PATH)
    total_md = sum(len(s.test_cases) for s in sections)

    if args.check:
        if not XLSX_PATH.exists():
            print(f"FAIL: Excel file does not exist: {XLSX_PATH}")
            sys.exit(1)

        diffs = check_sync(sections, XLSX_PATH)
        if diffs:
            print(f"FAIL: {len(diffs)} difference(s) between markdown ({total_md} TCs) and Excel:")
            for d in diffs:
                print(d)
            print(f"\nRun 'python scripts/sync_qa_test_cases.py --sync' to update.")
            sys.exit(1)
        else:
            print(f"OK: Excel is in sync with markdown ({total_md} test cases)")
            sys.exit(0)

    if args.sync:
        tracking = _load_existing_tracking(XLSX_PATH)
        total = _write_xlsx(sections, tracking, XLSX_PATH)
        print(f"Synced {total} test cases across {len(sections)} sections → {XLSX_PATH.name}")


if __name__ == "__main__":
    main()
